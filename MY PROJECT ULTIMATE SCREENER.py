import requests
from bs4 import BeautifulSoup
import time
import pandas as pd
from openpyxl import Workbook, load_workbook

#get_global()
#get_equity()
#get_fs()
#get_asian()
#get_commodities()
#get_forex()
#get_bonds()
#get_recommendation()

#ALL BASE IN INVESTING.COM

tech_stocks = ['CNVRG','GLO','TEL']
bluechip_stocks = ['AC', 'ACEN','AEV', 'AGI','ALI','AP','BDO','BPI','CNVRG','EMP','GLO','GTCAP','ICT','JFC','JGS','LTG','MBT','MEG','MER','MONDE','MPI','PGOLD','RLC','SECB','SM','SMC','SMPH','TEL','URC','WLCON']
bank_stocks = ['BDO', 'BPI', 'CHIB', 'EW', 'MBT', 'PNB', 'SECB', 'UBP']
oil_stocks = ['PXP(brent-oil)']
coal_stocks = ['SCC', 'DMC']
copper_stocks = ['AT', 'LC', 'PX']
gold_stocks = ['APX','PX','LC']
nickel_stocks = ['NIKL', 'FNI', 'MARC']
palm_stock = ['MONDE', 'URC', 'DNL']



############################################# NEED TO SO SUBSIDIARY

def indices_info(ind):
    url = 'https://www.investing.com/indices/'+ind
    r = requests.get(url)
    web = BeautifulSoup(r.text, 'html.parser')
    web = web.find('div', class_='text-xl flex items-end flex-wrap').text
    return web


def commodity_info(commodity):
    url = 'https://www.investing.com/commodities/'+commodity
    r = requests.get(url)
    web = BeautifulSoup(r.text, 'html.parser')
    web = web.find('div', class_='text-xl flex items-end flex-wrap').text
    return web
    

    
def bank_info(bank):
    url = 'https://www.investing.com/equities/' + bank
    r = requests.get(url)
    web = BeautifulSoup(r.text, 'html.parser')
    web = web.find('div', class_='text-xl flex items-end flex-wrap').text
    return web


    
    
def index_percentage(indexcode):
    url = 'https://www.investing.com/indices/' + indexcode
    r = requests.get(url)

    web = BeautifulSoup(r.text, 'html.parser')
    web = web.find('div', class_='text-xl flex items-end flex-wrap').text
    web = web[-7:-2]
    web = float(web)

    return web

def commodities_percentage(commoditiescode):
    url = 'https://www.investing.com/commodities/' + commoditiescode
    r = requests.get(url)

    web = BeautifulSoup(r.text, 'html.parser')
    web = web.find('div', class_='text-xl flex items-end flex-wrap').text
    web = web[-7:-2]
    web = float(web)

    return web

def coal_percentage(coalcode):
    url = 'https://www.investing.com/equities/'+coalcode+'-india-ltd-future'
    r = requests.get(url)

    web = BeautifulSoup(r.text, 'html.parser')
    web = web.find('div', class_='top bold inlineblock').text
    web = web[-7:-2]
    web = float(web)

    return web

def get_palm():
    palm_url = 'https://markets.businessinsider.com/commodities/palm-oil-price'
    palm_r = requests.get(palm_url)
    palm = BeautifulSoup(palm_r.text, 'html.parser')
    palm_current = palm.find('span', class_='price-section__current-value').text
    palm_points = palm.find('span', class_='price-section__absolute-value').text
    palm_percentage = float(palm.find('span', class_='price-section__relative-value').text[:-1])

    return palm_percentage


def bank_percentage(bank):
    url = 'https://www.investing.com/equities/' + bank
    r = requests.get(url)

    web = BeautifulSoup(r.text, 'html.parser')
    web = web.find('div', class_='text-xl flex items-end flex-wrap').text
    web = web[-7:-2]
    web = float(web)
    return web

def bonds_percentage(country,year):
    url = 'https://www.investing.com/rates-bonds/'+country+'-'+ year +'-year-bond-yield'
    r = requests.get(url)
    web = BeautifulSoup(r.text, 'html.parser')
    web = web.find('div', class_='top bold inlineblock').text
    web_price = float(web[:6])
    web_change = float(web[7:13])
    web_percentage = float(web[17:-2])

    return web_percentage

def forex_percentage():
    url = 'https://www.investing.com/currencies/usd-php'
    r = requests.get(url)
    web = BeautifulSoup(r.text, 'html.parser')
    web_price = web.find('span', class_='text-2xl').text
    web_change = web.find('div', class_='text-xl flex items-end flex-wrap').text
    web_changes = web_change[:6]
    web_percentage = float(web_change[7:-2])

    return web_percentage

bank = {bank_percentage('jp-morgan-chase') : 'JP MORGAN CHASE'}

bank_list_percentage = [bank_percentage('jp-morgan-chase'), bank_percentage('bank-of-america'), bank_percentage('wells-fargo'), bank_percentage('citigroup'), bank_percentage('us-bancorp'), bank_percentage('bb-t-corp'), bank_percentage('pnc-fin-serv'), bank_percentage('toronto-dominion-bank?cid=20605'), bank_percentage('capital-one'), bank_percentage('bk-of-ny')]



#############################################################################


def get_equity():
    print(f"INDICES: \nDow Jones: {indices_info('us-30')} \nNASDAQ: {indices_info('nasdaq-composite')} ")

    print('________________________________________\n')

    print(f"BANKS: \nJP MORGAN CHASE: {bank_info('jp-morgan-chase')} % \nBANK OF AMERICA: {bank_info('bank-of-america')} % \nWELLS FARGO: {bank_info('wells-fargo')} % \nCITIGROUP: {bank_info('citigroup')} % \nUS BANCORP: {bank_info('us-bancorp')} % \nTRUIST FINANCIAL CORP: {bank_info('bb-t-corp' )} % \nPNC FINANCIAL: {bank_info('pnc-fin-serv')} % \nTORONTO DOMINION BANK: {bank_info('toronto-dominion-bank?cid=20605')} % \nCAPITAL ONE FINANCIAL CORP: {bank_info('capital-one')} % \nBANK OF NEW YORK MELLON: {bank_info('bk-of-ny')} %\n\n")

def get_commodities():

    print(f"COMMODITIES: \nNICKEL :{commodity_info('nickel?cid=959208')}  \nGOLD: {commodity_info('gold')}  \nCopper: {commodity_info('copper')}  \nOil: {commodity_info('brent-oil')}  \nCoal: {coal_percentage('coal')} % \nPalm oil: {get_palm()} %")

    print('________________________________________\n')

def get_forex():
    url = 'https://www.investing.com/currencies/usd-php'
    r = requests.get(url)
    web = BeautifulSoup(r.text, 'html.parser')
    web_price = web.find('span', class_='text-2xl').text
    web_change = web.find('div', class_='text-xl flex items-end flex-wrap').text
    web_changes = web_change[:6]
    web_percentage = float(web_change[7:-2])
    print(f"USD/PHP Exchange Rate: {web_price} {web_changes} {web_percentage} %")
    print(f"####################### FOREX #######################\n")
    
    

def get_bonds():
    print(f"US 10 Year Bond Yield: {bonds_percentage('u.s.', '10')} %")
    print(f"US 5 Year Bond Yield: {bonds_percentage('u.s.', '5')} %")
    print(f"PH 10 Year Bond Yield: {bonds_percentage('philippines', '10')} %")
    print(f"PH 5 Year Bond Yield: {bonds_percentage('philippines', '5')} %")
    print(f"PH 1 Year Bond Yield: {bonds_percentage('philippines', '1')} %")
    print('________________________________________\n')
    
########################################################### RECOMMENDATION ###########################################################

def get_recommendation():
    print('__________POSSIBLE PLAYS BELOW_________\n')
        
    if index_percentage('us-30') >= 1.3:
        print(f"Down Jones (DJI): {index_percentage('us-30')} %\nShould look into:\n{bluechip_stocks}\n_________________________________\nIf you have {bluechip_stocks} you might want to sell your position today.\n________________________________________\n")
        

    elif index_percentage('us-30') <= -1.3:
        print(f"WATCHOUT for {bluechip_stocks}!! DJI is down {index_percentage('us-30')}% \nIf you do not have {bluechip_stocks} you might consider to buy today.\n________________________________________\n" )

    #################################
    if bonds_percentage('u.s.', '10') > 1.5:
        print(f"US 10 Year Treasury Yield Up {bonds_percentage('u.s.', '10')} %")
    
    elif bonds_percentage('u.s.', '10') < -1.5:
        print(f"US 10 Year Treasury Yield Down {bonds_percentage('u.s.', '10')} %")
    
    if bonds_percentage('u.s.', '5') > 1.5:
        print(f"US 5 Year Treasury Yield Up {bonds_percentage('u.s.', '5')} %")
    
    elif bonds_percentage('u.s.', '5') < -1.5:
        print(f"US 5 Year Treasury Yield Down {bonds_percentage('u.s.', '5')} %")
    
    if bonds_percentage('philippines', '10') > 1.5:
        print(f"Philippine 10 Year Treasury Yield Up {bonds_percentage('philippines', '10')} %")
        
    elif bonds_percentage('philippines', '10') < -1.5:
        print(f"Philippine 10 Year Treasury Yield Down {bonds_percentage('philippines', '10')} %")
    
    if bonds_percentage('philippines', '5') > 1.5:
        print(f"Philippine 5 Year Treasury Yield Up {bonds_percentage('philippines', '5')} %")
    
    elif bonds_percentage('philippines', '5') < -1.5:
        print(f"Philippine 5 Year Treasury Yield Down {bonds_percentage('philippines', '5')} %")
        
    if bonds_percentage('philippines', '1') > 1.5:
        print(f"Philippine 1 Year Treasury Yield Up {bonds_percentage('philippines', '1')} %")
    
    elif bonds_percentage('philippines', '1') < -1.5:
        print(f"Philippine 1 Year Treasury Yield Down {bonds_percentage('philippines', '1')} %")
    
    if commodities_percentage('nickel?cid=959208') > 1.5:
        print(f"COMMODITIES: \nNickel Futures(Commodity): {commodities_percentage('nickel?cid=959208')}% \nShould Look into:\n{nickel_stocks}\nIf you have {nickel_stocks} you might want to sell your position today.\nCharts Here: https://www.investing.com/commodities/nickel-streaming-chart?cid=959208\n________________________________________\n")
        

    elif commodities_percentage('nickel?cid=959208') <= -1.5:
        print(f"WATCHOUT for {nickel_stocks}!! Nickel is down {commodities_percentage('nickel?cid=959208')}% \nIf you do not have {nickel_stocks} you might consider to buy today.\nFinancial Reports for NIKL: https://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=625\nFinancial Reports for FNI: https://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=224\nFinancial Reports for MARC: https://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=175\n________________________________________\n" )

        
    if commodities_percentage('gold') > 1.5:
        print(f"COMMODITIES: \nGold Futures(Commodity): {commodities_percentage('gold')} \nShould look into:\n{gold_stocks}\nIf you have {gold_stocks} you might want to sell your position today.\n________________________________________\n")

    elif commodities_percentage('gold') <= -1.5:
        print(f"WATCHOUT for {gold_stocks}!! Gold is down {commodities_percentage('gold')}% \nIf you do not have {gold_stocks} you might consider to buy today.\nFinancial Reports for APX: https://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=178\nFinancial Reports for PX: https://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=137\nFinancial Reports for LC: https://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=98\n________________________________________\n" )

    if commodities_percentage('copper') > 1.5:
        print(f"COMMODITIES: \nCopper Futures(Commodity): {commodities_percentage('copper')} \nShould look into:\n{copper_stocks}\nIf you have {copper_stocks} you might want to sell your position today.\n________________________________________\n")

    elif commodities_percentage('copper') <= -1.5:
        print(f"WATCHOUT for {copper_stocks}!! Copper is down {commodities_percentage('copper')}% \nIf you do not have {copper_stocks} you might consider to buy today.\nFinancial Reports for AT: https://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=34\nFinancial Reports for LC: https://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=98\nhttps://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=137\n________________________________________\n" )


    if commodities_percentage('brent-oil') > 1.5:
        print(f"COMMODITIES: \nBrent-oil Futures(Commodity): {commodities_percentage('brent-oil')} %  \nShould look into:\n{oil_stocks}\nIf you have {oil_stocks} you might want to sell your position today.\n________________________________________\n")

    elif commodities_percentage('brent-oil') <= -1.5:
        print(f"WATCHOUT for {oil_stocks}!! Brent-oil is down {commodities_percentage('brent-oil')}% \nIf you do not have {oil_stocks} you might consider to buy today.\nFinancial Reports for PXP: https://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=628\n________________________________________\n" )

        
    if coal_percentage('coal') >= 1.5:
        print(f"COMMODITIES: \nCoal: {coal_percentage('coal')}  \nShould look into:\n{coal_stocks}\nIf you have {coal_stocks} you might want to sell your position today.\n________________________________________\n")
    
    elif coal_percentage('coal') <= -1.5:
        print(f"WATCHOUT for {coal_stocks}!! COAL is down {coal_percentage('coal')}% \nIf you do not have {coal_stocks} you might consider to buy today.\nFinancial Reports for SCC: https://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=157\nFinancial Reports for DMC: https://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=188\n________________________________________\n" )

    if get_palm() > 1.5:
        print(f"WATCHOUT for Palm Oil!! Palm oil is Up {get_palm()} %\nUp Price is negative to {palm_stock}\nYou might want to consider buying today\nFinancial Reports for MONDE:https://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=682\nFinancial Reports for URC: https://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=124\nFinancial Reports for DNL: https://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=639\n_________________________________ ")

    elif get_palm() < -1.5:
        print(f"COMMODITIES: Palm Oil: Down {get_palm()} %\nDown Price is positive to {palm_stock}\nYou might want to consider selling today\nFinancial Reports for MONDE:https://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=682\nFinancial Reports for URC: https://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=124\nFinancial Reports for DNL: https://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=639\n_________________________________")
        
    bankpositive = []
    banknegative = []
    for bank_list in bank_list_percentage:
        if bank_list > 1.3:
            bankpositive.append(bank_list)
        
        elif bank_list <= -1.3:
            banknegative.append(bank_list)
        
    if len(bankpositive) >= 4:
        print(f"There are 5 or more positive in Banks \nShould look into:\n {bank_stocks}\nIf you have {bank_stocks} you might want to sell your position today.\n________________________________________ \n" )

    elif len(banknegative) >= 4:
        print(f"WATCHOUT for {bank_stocks} BANKS are down {len(banknegative)} items.\nIf you do not have {bank_stocks} you might consider to buy today.")
        print(f"Financial Reports for BDO: https://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=260")
        print(f"Financial Reports for BPI: https://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=234")
        print(f"Financial Reports for CHIB: https://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=184")
        print(f"Financial Reports for EW: https://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=634")
        print(f"Financial Reports for MBT: https://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=128")
        print(f"Financial Reports for PNB: https://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=139")
        print(f"Financial Reports for SECB: https://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=32")
        print(f"Financial Reports for UBP: https://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=167")
        print('\n________________________________________\n')

def get_asian():
    print(f"SOME ASIAN MARKET AS OF THE MOMENT: \n\nJAPAN NIKKEI*: {indices_info('japan-ni225')}\nCHINA SHANGHAI(SSEC): {indices_info('shanghai-composite')}\nCHINA SZSE: {indices_info('szse-component')}\nCHINA SSE 100: {indices_info('sse-100')}\nHONGKONG HANG SENG*: {indices_info('hang-sen-40')}\nKOREA KOSPI*:{indices_info('kospi')}\nINDIA BSE SENSEX: {indices_info('sensex')}\nTAIWAN: {indices_info('taiwan-weighted')}\nAUSTRALIA ASX 200: {indices_info('aus-200')}\nPAKISTAN KARACHI 100: {indices_info('karachi-100')}")
    print('##############################################################################################')

def get_fs(url,stock):
    # GETTING DATA RETRIEVING THROUGH WEBSCRAPING
    geturl = url
    r = requests.get(geturl)

    
    stock = pd.read_html(r.text)

    #ASSIGNING TABLES BASED ON FINANCIAL STATEMENT
    stock_annual_balance_sheet = stock[0]
    stock_annual_balance_sheet['Computed net'] = (stock_annual_balance_sheet['Current Year'] - stock_annual_balance_sheet['Previous Year']) / stock_annual_balance_sheet['Previous Year'] * 100

    stock_annual_income_statement = stock[1]
    stock_annual_income_statement['Computed net'] = (stock_annual_income_statement['Current Year'] - stock_annual_income_statement['Previous Year']) / stock_annual_income_statement['Previous Year'] * 100

    stock_quarterly_income_statement = stock[3]
    stock_quarterly_income_statement['3 Months Computed Net'] = (stock_quarterly_income_statement['Current Year (3 Months)'] - stock_quarterly_income_statement['Previous Year (3 Months)']) / stock_quarterly_income_statement['Previous Year (3 Months)'] * 100
    stock_quarterly_income_statement['YTD Computed Net'] = (stock_quarterly_income_statement['Current Year-To-Date'] - stock_quarterly_income_statement['Previous Year-To-Date']) / stock_quarterly_income_statement['Previous Year-To-Date']* 100

    #DATA CLEANSING , DROPPING UNRELATED COLUMN
    stock_annual_balance_sheet = stock_annual_balance_sheet.drop(['Current Year', 'Previous Year'], axis=1)
    stock_annual_income_statement = stock_annual_income_statement.drop(['Current Year', 'Previous Year'], axis=1)
    stock_quarterly_income_statement = stock_quarterly_income_statement.drop(['Current Year (3 Months)','Previous Year (3 Months)','Current Year-To-Date','Previous Year-To-Date'], axis=1)

    print('############# Annual Financial Reports #############')
    print(stock_annual_balance_sheet)
    print(stock_annual_income_statement)
    print('############# Quarterly Financial Reports #############')
    print(stock_quarterly_income_statement)
    

def get_global():
    get_commodities()
    get_forex()
   # get_bonds()
    get_equity()




def update_excel():
    wb = load_workbook('Global Data.xlsx')
    ws = wb.active


    ws['E6'] = commodities_percentage('nickel?cid=959208')
    ws['E7'] = commodities_percentage('gold')
    ws['E8'] = commodities_percentage('copper')
    ws['E9'] = commodities_percentage('brent-oil')
    ws['E10'] = coal_percentage('coal')
    ws['E11'] = get_palm()
    ws['E14'] = forex_percentage()
   # ws['E17'] = bonds_percentage('u.s.', '10')
   # ws['E18'] = bonds_percentage('u.s.', '5')
  #  ws['E19'] = bonds_percentage('philippines', '10')
   # ws['E20'] = bonds_percentage('philippines', '5')
    ws['H6'] = index_percentage('us-30')
    ws['H7'] = index_percentage('nasdaq-composite')
    ws['H11'] = index_percentage('japan-ni225')
    ws['H12'] = index_percentage('shanghai-composite')
    ws['H13'] = index_percentage('szse-component')
    ws['H14'] = index_percentage('sse-100')
    ws['H15'] = index_percentage('hang-sen-40')
    ws['H16'] = index_percentage('kospi')
    ws['H17'] = index_percentage('sensex')
    ws['H18'] = index_percentage('taiwan-weighted')
    ws['H19'] = index_percentage('aus-200')
    ws['H20'] = index_percentage('karachi-100')
    
    wb.save('Global Data.xlsx')
    wb.close()


get_global()

#while True:
#    get_asian()
#    time.sleep(300)

#get_commodities()
#get_forex()
#get_bonds()
#get_equity()

#update_excel()

get_recommendation()

#get_fs('https://edge.pse.com.ph/companyPage/financial_reports_view.do?cmpy_id=639', 'dnl')


